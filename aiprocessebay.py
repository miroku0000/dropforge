import sys
import glob
import pandas as pd
import openpyxl
import requests
import json
import time
import shutil
import os
import hashlib
import threading
import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from bs4 import BeautifulSoup

from ebay_utils import get_ebay_description


# --- Configuration ---
HEADER_ROW = 4  # Row number where the main header starts (1-based index)
LOG_FILE = "update_log.txt"
PROMPT_LOG_FILE = None  # Dynamically set per file
INVALID_UPDATE_LOG_FILE = None  # Dynamically set per file
CACHE_DIR = None  # Dynamically set
FORCE_REFRESH_IDS = set()  # Set of Item IDs to force refresh AI query for
FORCE_ALL = False  # Set to True via command line arg to force all
LLM_TIME_LOG = os.path.join("..", "data", "llmtime.txt")
MAX_VALUE_LENGTH = 65  # Max character length for each item specific value
MAX_ITEM_SPECIFICS = 45  # Max number of item specifics per listing

REQUIRED_FIELDS = []  # Example: ['Brand', 'Size Type', 'Size (Men\'s)']
PREFERRED_FIELDS = []  # Example: ['Color', 'Style', 'Material']

model_timings = {}
progress_lock = threading.Lock()
progress_counter = {"current": 0, "total": 0}


def parse_args():
    parser = argparse.ArgumentParser(
        description="Process eBay listings and update item specifics using AI."
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Force refresh ALL AI queries (delete cache).",
    )
    parser.add_argument(
        "--refresh", nargs="*", help="Force refresh AI query for specific Item IDs."
    )
    parser.add_argument(
        "--model_name",
        type=str,
        default="gemma3:4b",
        help="Specify the LLM model name to use.",
    )
    return parser.parse_args()


def update_llm_timing_log(model, elapsed_time):
    if not os.path.exists(LLM_TIME_LOG):
        os.makedirs(os.path.dirname(LLM_TIME_LOG), exist_ok=True)
        with open(LLM_TIME_LOG, "a", encoding="utf-8") as f:
            pass
    model_stats = {}
    if os.path.exists(LLM_TIME_LOG):
        try:
            with open(LLM_TIME_LOG, "r", encoding="utf-8") as f:
                for line in f:
                    parts = line.strip().split(":")
                    if len(parts) == 2:
                        model_name, stats_part = parts
                        stats = stats_part.split("(")[0].strip()
                        if stats:
                            try:
                                times = list(map(float, stats.split(",")))
                                model_stats[model_name] = times
                            except ValueError:
                                print(
                                    f"Warning: Could not parse times for model {model_name} in {LLM_TIME_LOG}"
                                )
                                model_stats[model_name] = []
                        else:
                            model_stats[model_name] = []
        except Exception as e:
            print(f"Error reading LLM time log: {e}")
            model_stats = {m: [] for m in model_stats}
    times = model_stats.get(model, [])
    times.append(elapsed_time)
    model_stats[model] = times
    try:
        with open(LLM_TIME_LOG, "w", encoding="utf-8") as f:
            for m, t_list in model_stats.items():
                if t_list:
                    avg = sum(t_list) / len(t_list)
                    f.write(f"{m}: {','.join(map(str, t_list))} (avg: {avg:.2f})\n")
                else:
                    f.write(f"{m}: \n")
    except Exception as e:
        print(f"Error writing LLM time log: {e}")


def extract_constraints_from_aspects(file_path):
    try:
        aspects_df = pd.read_excel(
            file_path, sheet_name="Aspects", dtype=str, header=None, skiprows=1
        ).fillna("")
        num_cols = aspects_df.shape[1]
        base_cols = ["Category ID", "Aspect Name", "Unused_1", "Unused_2", "Level"]
        value_cols = [f"Value_{i}" for i in range(5, num_cols)]
        aspects_df.columns = base_cols + value_cols
        constraints = {}
        for _, row in aspects_df.iterrows():
            aspect_name = row["Aspect Name"].strip()
            if not aspect_name:
                continue
            allowed_values = [
                str(row[col]).strip()
                for col in aspects_df.columns[5:]
                if str(row[col]).strip()
            ]
            if allowed_values:
                constraints[aspect_name] = {"type": "list", "values": allowed_values}
        return constraints
    except Exception as e:
        print(f"Error reading or processing 'Aspects' sheet in {file_path}: {e}")
        print("Proceeding without constraints.")
        return {}


def fetch_ebay_listing(ebay_item_id, max_retries=3):
    description = get_ebay_description(ebay_item_id)
    # print(description)
    return description


def log_update(item_id, field, old_value, new_value):
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as log_file:
            log_file.write(
                f"Item ID: {item_id}, Field: {field}, Old: '{old_value}', New: '{new_value}'\n"
            )
    except Exception as e:
        print(f"Error writing to log file {LOG_FILE}: {e}")


def log_invalid_update(item_id, field, proposed_value, reason):
    if INVALID_UPDATE_LOG_FILE:
        try:
            with open(INVALID_UPDATE_LOG_FILE, "a", encoding="utf-8") as invalid_log:
                invalid_log.write(
                    f"Item ID: {item_id}, Field: {field}, Proposed: '{proposed_value}', Reason: {reason}\n"
                )
        except Exception as e:
            print(
                f"Error writing to invalid update log file {INVALID_UPDATE_LOG_FILE}: {e}"
            )


def log_prompt_and_response(item_id, prompt, response):
    global PROMPT_LOG_FILE
    if PROMPT_LOG_FILE:
        try:
            with open(PROMPT_LOG_FILE, "a", encoding="utf-8") as prompt_log:
                prompt_log.write(
                    f"--- Item ID: {item_id} ---\nPrompt:\n{prompt}\n\nResponse:\n{response}\n\n---\n\n"
                )
        except Exception as e:
            print(f"Error writing to prompt log file {PROMPT_LOG_FILE}: {e}")


def cache_lookup(prompt):
    if not CACHE_DIR:
        return None
    prompt_hash = hashlib.sha256(prompt.encode("utf-8")).hexdigest()
    cache_path = os.path.join(CACHE_DIR, f"{prompt_hash}.json")
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError) as e:
            print(f"Error reading cache file {cache_path}: {e}. Ignoring cache.")
            return None
    return None


def cache_store(prompt, response_text):
    if not CACHE_DIR:
        return
    prompt_hash = hashlib.sha256(prompt.encode("utf-8")).hexdigest()
    cache_path = os.path.join(CACHE_DIR, f"{prompt_hash}.json")
    try:
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump({"response": response_text}, f)
    except OSError as e:
        print(f"Error writing cache file {cache_path}: {e}")


def limit_item_specifics(
    item_specifics, required, preferred, max_count=MAX_ITEM_SPECIFICS
):
    current_specifics = {k: v for k, v in item_specifics.items() if str(v).strip()}
    current_count = len(current_specifics)
    if current_count <= max_count:
        return item_specifics
    num_to_remove = current_count - max_count
    removable_non_priority = [
        field
        for field, value in current_specifics.items()
        if field not in required and field not in preferred and str(value).strip()
    ]
    removed_count = 0
    for field in removable_non_priority:
        if removed_count < num_to_remove:
            item_specifics[field] = ""
            removed_count += 1
        else:
            break
    if (current_count - removed_count) > max_count:
        num_still_to_remove = (current_count - removed_count) - max_count
        removable_preferred = [
            field
            for field, value in current_specifics.items()
            if field in preferred
            and field not in required
            and str(item_specifics.get(field, "")).strip()
        ]
        for field in removable_preferred:
            if removed_count < num_to_remove:
                item_specifics[field] = ""
                removed_count += 1
            else:
                break
    final_count = len([v for v in item_specifics.values() if str(v).strip()])
    if final_count > max_count:
        print(
            f"Warning: Item specifics count ({final_count}) still exceeds max ({max_count}) after attempting removal. This might happen if there are more than {max_count} required fields."
        )
    return item_specifics


def validate_ai_response(response, constraints, existing_row, item_id):
    validated = {}
    if not isinstance(response, dict):
        print(f"Warning: AI response for {item_id} is not a dictionary: {response}")
        return {}
    for field, value in response.items():
        if field is None or value is None:
            continue
        if isinstance(value, list):
            processed_values = [
                str(v)[:MAX_VALUE_LENGTH] for v in value if str(v).strip()
            ]
            proposed_value = "||".join(processed_values) if processed_values else ""
        else:
            proposed_value = str(value)[:MAX_VALUE_LENGTH]
        if not proposed_value.strip():
            continue
        if field in constraints:
            constraint_info = constraints[field]
            allowed_values = constraint_info.get("values", [])
            is_valid_constrained_value = False
            if isinstance(value, list):
                original_parts = [str(v) for v in value if str(v).strip()]
                valid_parts = [
                    part for part in original_parts if part in allowed_values
                ]
                if len(valid_parts) == len(original_parts) and original_parts:
                    validated_value_joined = "||".join(
                        [p[:MAX_VALUE_LENGTH] for p in valid_parts]
                    )
                    validated[field] = validated_value_joined[:MAX_VALUE_LENGTH]
                    is_valid_constrained_value = True
            else:
                if proposed_value in allowed_values:
                    validated[field] = proposed_value
                    is_valid_constrained_value = True
            if not is_valid_constrained_value:
                existing_value = str(existing_row.get(field, "")).strip()
                if existing_value.lower() in [
                    "nan",
                    "null",
                    "undefined",
                    "",
                    "does not apply",
                    "none",
                    "n/a",
                ]:
                    existing_value = ""
                if existing_value == "":
                    log_invalid_update(
                        item_id,
                        field,
                        proposed_value,
                        f"Proposed value ('{proposed_value}') not in allowed list, and existing value was blank. Skipping update for this field.",
                    )
                else:
                    log_invalid_update(
                        item_id,
                        field,
                        proposed_value,
                        f"Proposed value ('{proposed_value}') not in allowed list and existing value ('{existing_value}') is not blank. Skipping update.",
                    )
                continue
        else:
            validated[field] = proposed_value
    return {k: v for k, v in validated.items() if str(v).strip()}


def update_openpyxl_sheet_from_df(ws, df, start_row=HEADER_ROW):
    max_df_row = start_row + len(df)
    for row_idx in range(start_row + 1, max(ws.max_row + 1, max_df_row + 1)):
        for col_idx in range(1, df.shape[1] + 1):
            ws.cell(row=row_idx, column=col_idx).value = None
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, value in enumerate(row, start=1):
            if pd.isna(value):
                cell_value = None
            else:
                cell_value = str(value)
            ws.cell(row=r_idx, column=c_idx, value=cell_value)


def process_item(args, constraints, model_name):
    start_item_time = time.time()
    idx, item_id, row, missing_fields = args
    with progress_lock:
        progress_counter["current"] += 1
        current = progress_counter["current"]
        total = progress_counter["total"]
        print(f"Processing {current}/{total} (Item ID: {item_id})...")
    validated_specifics = {}
    try:
        description = fetch_ebay_listing(item_id)
        if not description:
            print(
                f"Warning: No description found for Item ID {item_id}. AI may have limited context."
            )
        prompt = f"""Analyze the eBay listing description below for item {item_id}.
Suggest values for these missing item specifics: {', '.join(missing_fields)}.
Provide your answer STRICTLY as a JSON object where keys are the field names and values are the suggested specifics.
Each specific value MUST NOT exceed {MAX_VALUE_LENGTH} characters.
If you cannot determine a value for a field based on the description, omit that field from your JSON response.
Do not invent fields not listed above. Do not add explanations outside the JSON structure.
Description:
---
{description[:3000]}
---
End Description. Provide JSON response:"""
        use_cache = not FORCE_ALL and item_id not in FORCE_REFRESH_IDS
        cached = cache_lookup(prompt) if use_cache else None
        result_text = None
        if cached:
            print(f"Cache hit for item {item_id}")
            result_text = cached.get("response")
            if not result_text:
                print(
                    f"Warning: Cache entry found but empty for {item_id}. Refetching."
                )
                cached = None
        if not cached:
            print(f"Calling AI ({model_name}) for item {item_id}...")
            ai_start_time = time.time()
            retries = 2
            delay = 5
            response_data = None
            for attempt in range(retries + 1):
                try:
                    response = requests.post(
                        "http://localhost:11434/api/generate",
                        json={
                            "model": model_name,
                            "prompt": prompt,
                            "format": "json",
                            "stream": False,
                            "options": {"temperature": 0.2},
                        },
                        timeout=480,
                    )
                    response.raise_for_status()
                    response_data = response.json()
                    result_text = response_data.get("response", "").strip()
                    if result_text.startswith("{") and result_text.endswith("}"):
                        break
                    else:
                        print(
                            f"Warning: AI response for {item_id} (attempt {attempt+1}) not valid JSON: {result_text[:100]}..."
                        )
                        break
                except requests.exceptions.ReadTimeout:
                    print(
                        f"AI Timeout on attempt {attempt+1} for {item_id}. Retrying in {delay}s..."
                    )
                    time.sleep(delay)
                    delay *= 2
                except requests.exceptions.RequestException as e:
                    print(f"AI Request Error on attempt {attempt+1} for {item_id}: {e}")
                    time.sleep(delay)
                    delay *= 2
                except Exception as e:
                    print(
                        f"Unexpected Error during AI call for {item_id} (attempt {attempt+1}): {e}"
                    )
                    time.sleep(delay)
                    delay *= 2
            ai_elapsed_time = time.time() - ai_start_time
            if response_data:
                update_llm_timing_log(model_name, ai_elapsed_time)
            print(f"AI call for {item_id} took {ai_elapsed_time:.2f}s")
            if not result_text:
                print(
                    f"Error: AI ({model_name}) did not return a response body for item {item_id}."
                )
                return (idx, None)
            else:
                print(
                    f"Error: AI ({model_name}) failed for item {item_id} after retries."
                )
                return (idx, None)
        if result_text:
            cache_store(prompt, result_text)
        log_prompt_and_response(
            item_id, prompt, result_text if result_text else "[No Response]"
        )
        if not result_text:
            print(f"No result text received from AI or cache for {item_id}.")
            return (idx, {})
        parsed_response = {}
        try:
            json_start = result_text.find("{")
            json_end = result_text.rfind("}") + 1
            if json_start != -1 and json_end > json_start:
                json_str = result_text[json_start:json_end]
                parsed_response = json.loads(json_str)
                if not isinstance(parsed_response, dict):
                    print(
                        f"Warning: Parsed JSON for {item_id} is not a dictionary: {type(parsed_response)}"
                    )
                    parsed_response = {}
            else:
                print(
                    f"Warning: Could not find JSON object markers '{{' and '}}' in AI response for {item_id}."
                )
        except json.JSONDecodeError as e:
            print(f"Error decoding AI JSON response for {item_id}: {e}")
            print(f"Raw Response Snippet: {result_text[:500]}")
            with open(f"debug_ai_response_{item_id}.txt", "w", encoding="utf-8") as f:
                f.write(result_text)
            parsed_response = {}
        validated_specifics = validate_ai_response(
            parsed_response, constraints, row, item_id
        )
    except Exception as e:
        print(f"!!! Critical Error processing item {item_id} (Index {idx}): {e}")
        import traceback

        traceback.print_exc()
        return (idx, None)
    finally:
        end_item_time = time.time()
        print(
            f"Finished Item ID: {item_id}. Time: {end_item_time - start_item_time:.2f}s. Found {len(validated_specifics)} valid fields."
        )
    return (idx, validated_specifics)


def main(model_name):
    global PROMPT_LOG_FILE, INVALID_UPDATE_LOG_FILE, CACHE_DIR, FORCE_REFRESH_IDS, FORCE_ALL
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, "..", "data")
    CACHE_DIR = os.path.join(output_dir, "aicache")
    CSV_OUTPUT_DIR = os.path.join(output_dir, "aicsvtoupload")
    PROCESSED_DIR = os.path.join(output_dir, "processed")
    os.makedirs(CACHE_DIR, exist_ok=True)
    os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)
    os.makedirs(PROCESSED_DIR, exist_ok=True)

    args = parse_args()
    if args.force:
        print("--- Force refreshing ALL AI queries! ---")
        FORCE_ALL = True
        if os.path.exists(CACHE_DIR):
            print(f"--- Deleting cache directory: {CACHE_DIR} ---")
            shutil.rmtree(CACHE_DIR)
        os.makedirs(CACHE_DIR)
    if args.refresh:
        FORCE_REFRESH_IDS = set(args.refresh)
        print(f"--- Force refreshing AI query for Item IDs: {FORCE_REFRESH_IDS} ---")
    model_name = args.model_name

    downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
    file_pattern = os.path.join(
        downloads_path, "eBay-Active-Listings-Item-Specifics-*.xlsx"
    )
    files = glob.glob(file_pattern)
    if not files:
        print(f"No files found matching pattern: {file_pattern}")
        return

    for file_path in files:
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        PROMPT_LOG_FILE = os.path.join(output_dir, f"{base_name}_prompt_log.txt")
        INVALID_UPDATE_LOG_FILE = os.path.join(
            output_dir, f"{base_name}_invalid_update_log.txt"
        )
        output_csv_path = os.path.join(CSV_OUTPUT_DIR, f"{base_name}_processed.csv")
        temp_excel_path = os.path.join(PROCESSED_DIR, f"{base_name}_temp.xlsx")
        print(f"\n>>> Processing file: {file_path}")
        try:
            print("Loading workbook...")
            workbook = openpyxl.load_workbook(file_path)
            if "Listings" not in workbook.sheetnames:
                print(f"Error: 'Listings' sheet not found in {file_path}. Skipping.")
                continue
            ws = workbook["Listings"]
            metadata = []
            for i in range(1, HEADER_ROW):
                row_data = [
                    str(cell.value) if cell.value is not None else "" for cell in ws[i]
                ]
                metadata.append(row_data)
            print("Reading Listings sheet into DataFrame...")
            df = pd.read_excel(
                file_path, sheet_name="Listings", dtype=str, header=HEADER_ROW - 1
            ).fillna("")
            if "Item ID" not in df.columns:
                print(
                    f"Error: 'Item ID' column not found in {file_path} (expected at header row {HEADER_ROW}). Skipping."
                )
                continue
            print("Extracting constraints from Aspects sheet...")
            constraints = {}
            if "Aspects" in workbook.sheetnames:
                constraints = extract_constraints_from_aspects(file_path)
                print(f"Found {len(constraints)} constraints.")
            else:
                print(
                    "Warning: 'Aspects' sheet not found. Proceeding without constraints."
                )
            print("Identifying items needing AI processing...")
            items_to_process = []
            blue_columns = [col for col in df.columns if col.startswith("C:")]
            if not blue_columns:
                print(
                    "Warning: No columns starting with 'C:' found. Assuming these are the item specifics columns."
                )
                potential_specific_cols = [
                    col
                    for col in df.columns
                    if col not in ["Item ID", "Listing Title", "Category ID"]
                ]
                blue_columns = potential_specific_cols
            if not blue_columns:
                print(
                    "Error: Could not identify any potential item specific columns. Skipping file."
                )
                continue
            for idx, row in df.iterrows():
                item_id_val = row.get("Item ID")
                if not item_id_val or str(item_id_val).strip().lower() in [
                    "",
                    "none",
                    "nan",
                ]:
                    print(
                        f"Skipping row {idx+HEADER_ROW+1} due to missing or invalid Item ID."
                    )
                    continue
                item_id = str(item_id_val).strip()
                missing_fields = []
                for field in blue_columns:
                    field_value = row.get(field, "")
                    if str(field_value).strip().lower() in [
                        "",
                        "does not apply",
                        "n/a",
                        "none",
                        "nan",
                    ]:
                        missing_fields.append(field)
                if missing_fields:
                    items_to_process.append((idx, item_id, row.copy(), missing_fields))
            print(
                f"Found {len(items_to_process)} items potentially needing AI suggestions."
            )
            if not items_to_process:
                print("No items require processing in this file.")
                processed_file_path = os.path.join(
                    PROCESSED_DIR, os.path.basename(file_path)
                )
                print(f"Moving original file to processed: {processed_file_path}")
                shutil.move(file_path, processed_file_path)
                continue
            progress_counter["total"] = len(items_to_process)
            progress_counter["current"] = 0
            results = []
            max_workers = 1
            print(f"Processing items using AI (max_workers={max_workers})...")
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {
                    executor.submit(
                        process_item, item_args, constraints, model_name
                    ): item_args
                    for item_args in items_to_process
                }
                for future in as_completed(futures):
                    item_args = futures[future]
                    idx, item_id, _, _ = item_args
                    try:
                        result_idx, validated_specifics = future.result()
                        if validated_specifics is not None:
                            results.append((result_idx, validated_specifics))
                        else:
                            print(
                                f"Item {item_id} (Index {idx}) failed processing critically."
                            )
                    except Exception as exc:
                        print(
                            f"Item {item_id} (Index {idx}) generated an exception: {exc}"
                        )
                        results.append((idx, {}))
            print("AI Processing complete.")
            print("Updating DataFrame with AI results...")
            updates_made = 0
            for idx, validated_specifics in results:
                if validated_specifics:
                    item_id = df.loc[idx, "Item ID"]
                    current_row_specifics = {}
                    for col in blue_columns:
                        if col in df.columns:
                            current_val = df.loc[idx, col]
                            if str(current_val).strip() and str(
                                current_val
                            ).strip().lower() not in ["nan", "none"]:
                                current_row_specifics[col] = current_val
                    merged_specifics = current_row_specifics.copy()
                    merged_specifics.update(validated_specifics)
                    limited_specifics = limit_item_specifics(
                        merged_specifics,
                        REQUIRED_FIELDS,
                        PREFERRED_FIELDS,
                        MAX_ITEM_SPECIFICS,
                    )
                    for field, value in limited_specifics.items():
                        if field in df.columns:
                            old_value = df.loc[idx, field]
                            new_value = str(value).strip()
                            if str(old_value).strip() != new_value:
                                if not new_value and str(old_value).strip():
                                    log_update(
                                        item_id,
                                        field,
                                        old_value,
                                        "[REMOVED BY LIMITER]",
                                    )
                                elif new_value:
                                    log_update(item_id, field, old_value, new_value)
                            df.loc[idx, field] = new_value
                            updates_made += 1
            print(f"Applied {updates_made} updates to the DataFrame.")
            print(f"Saving processed data to CSV: {output_csv_path}")
            try:
                with open(output_csv_path, "w", encoding="utf-8", newline="") as f:
                    for meta_row in metadata:
                        f.write(",".join(map(str, meta_row)) + "\n")
                    f.write(",".join(df.columns) + "\n")
                    df.to_csv(f, index=False, header=False, lineterminator="\n")
                print("CSV saved successfully.")
            except Exception as e:
                print(f"Error saving output files: {e}")
            try:
                processed_file_path = os.path.join(
                    PROCESSED_DIR, os.path.basename(file_path)
                )
                print(
                    f"Moving original file to processed directory: {processed_file_path}"
                )
                shutil.move(file_path, processed_file_path)
            except Exception as e:
                print(f"Error moving processed file {os.path.basename(file_path)}: {e}")
        except FileNotFoundError:
            print(f"Error: File not found {file_path}. Skipping.")
        except Exception as e:
            print(f"An unexpected error occurred processing file {file_path}: {e}")
            import traceback

            traceback.print_exc()
        finally:
            if "workbook" in locals() and workbook:
                try:
                    workbook.close()
                except Exception as e:
                    print(f"Minor error closing workbook: {e}")
        print("\n--- Script Finished ---")


if __name__ == "__main__":
    args = parse_args()
    main(args.model_name)
