from openpyxl import load_workbook
from pathlib import Path
import os
import glob
import json
import wmi
import time
# Initializing the wmi constructor

username = os.getenv("USERNAME")

def getDefaultkey(key):
	match key:
		case "Year Manufactured":
			return ""
		case "Custom Bundle":
			return "No"
		case "California Prop 65 Warning":
			return "Does Not Apply"
		case "Country/Region of Manufacture":
			return "Unknown"
		case "Personalization Instructions":
			return "Does Not Apply"
		case "Personalize":
			return "No"
		case "Handmade":
			return "No"
		case "Vintage":
			return "No"
		case "Antique":
			return "No"
		case "Recipient":
			return "Multipurpose"
		case "Department":
			return "Adults"
		case "Origin":
			return "Unknown"
		case "Pattern":
			return "Solid"
		case "Theme":
			return "Art"
		case "Style":
			return "Classicism"
		case default:
			return "See Listing"

def isRunning(s):
	# Iterating through all the running processes
	f = wmi.WMI()
	count=0
	for process in f.Win32_Process():	
		# Displaying the P_ID and P_Name of the process
		if process.CommandLine and s in process.CommandLine and "python.exe" in process.CommandLine:
			count=count+1
			print(str(process.ProcessId) +" " + process.CommandLine)
			if count>1:
				return True
	return False


	

def getItemSpecificsFromFile(id):
	print("debug getItemSpecificsFromFile called with id = " + id)
	myvars = {}
	with open("D:\\zikprocessor\\data\\itemspecifics\\" + id,  encoding="utf-8", errors="ignore") as myfile:
		for line in myfile:
			name, var = line.partition(":")[::2]		
			myvars[name.strip()] = var.strip()
		# 	print("debug key= |"+ name +"|")
		# 	print("debugvalue= |"+ var +"|")
		return myvars

#wb = load_workbook(filename = 'C:\\Users\\mirok\\Downloads\\eBay-Active-Listings-Item-Specifics-Feb-18-2023-12_19_52-0700-1384784468.xlsx')


def processFile(filename):
	wb = load_workbook(filename = filename)
	ws = wb['Listings']
	wis=wb['Aspects']
	# Get required item specifics
	istofillin={}
	for row in range(3, wis.max_row+1):
		category = wis.cell(row, 1).value
		level = wis.cell(row, 5).value
		aspectname= wis.cell(row, 2).value
		print(str(category))
		print(level)
		print(aspectname)
		if aspectname:
			aspectname=aspectname.replace("C:","")
		if not category in istofillin:
			istofillin[category]=[]
		if level and ("REQUIRED" in level or "PREFERRED" in level):
			istofillin[category].append(aspectname)
	print(json.dumps(istofillin))
	columnforis={}
	for row in ws.iter_rows(min_row=4, max_row=4):
		c=1
		for cell in row:
			if cell.value:
				specific=cell.value.replace("C:","").strip()
				print(specific)
				columnforis[specific]=c
			c=c+1
	for row in range(5, ws.max_row+1):
		id = ws.cell(row, 2).value
		if id:
			category = ws.cell(row, 1).value
			path = Path("D:\\zikprocessor\\data\\itemspecifics\\" + str(id))
			print(path)
			if not path.is_file():
				os.system("getitemspecificsfromebay.py " + id + " 1> D:\\zikprocessor\\data\\itemspecifics\\" + id)
			specifics=getItemSpecificsFromFile(id)
			for specname in columnforis:
				current = ws.cell(row, columnforis[specname]).value
				if not current:
					current=""
				# if we have a specific who we know the value of, and is currently not filled in
				if specname in specifics and (not current or "Does Not Apply" in current or "See Listing" in current or "None" in current):
					if specname in ["Features","Application", "Model","Theme", "Style", "Size", "Surface Finish", "Material", "Occasion", "Color"] and len(specifics[specname])>64:
						specifics[specname]=specifics[specname][0:64].replace('"',"")
					if "None" in str(specifics[specname]) or not specifics[specname]:
						specifics[specname]=getDefaultkey(specifics[specname])
					ws.cell(row, columnforis[specname], specifics[specname])
				else:
					if not current and category in istofillin and istofillin[category] and specname in istofillin[category]:					
						ws.cell(row, columnforis[specname], "See Listing")
				after = ws.cell(row, columnforis[specname]).value
				if after and current != after and after is not None: 	
					print ("bef " + str(id) + " " + specname + " : " + str(current) )
					print ("aft " + str(id) + " " + specname + " : |" + str(ws.cell(row, columnforis[specname]).value ) +"|")			
				if current and not after:
					print ("bef " + str(id) + " " + specname + " : " + str(current) )
					print("No after!!!!")
					input("Press Enter to continue...")
	outputdir="D:\\zikprocessor\\data\\ebayitemspecifics\\"
	outputfilename=outputdir+filename.split("\\")[-1]
	print(outputfilename)
	wb.save(outputfilename)

if  not isRunning("D:\\zikprocessor\\src\\testspreadsheet.py"):
	filenames = glob.glob("C:\\Users\\" + username +"\\Downloads\\eBay-Active-Listings-Item-Specifics*.xlsx")
	while filenames and len(filenames)>0:
		i=1
		l=len(filenames)
		for filename in filenames:
			print(str(i) + " of " + str(l) +" files :" + filename)
			processFile(filename)
			os.remove(filename) 
			i=i+1
			os.system("exceltocsv.py")
		os.system("exceltocsv.py")
		l=len(filenames)
		filenames = glob.glob("C:\\Users\\" + username +"\\Downloads\\eBay-Active-Listings-Item-Specifics*.xlsx")
else:
	print("testspreadsheet.py is already running...")
