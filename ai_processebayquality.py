import os
import glob
import re
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# === Settings ===
search_dir = "C:/Users/mirok/Downloads"
report_pattern = "Listing quality report for*.xlsx"
history_file = os.path.join(search_dir, "category_ranking_history.xlsx")
output_csv = os.path.join(search_dir, "deletebasedonquality.csv")

# === Step 1: Get All Matching Report Files ===
report_files = glob.glob(os.path.join(search_dir, report_pattern))
if not report_files:
    raise FileNotFoundError("No listing quality reports found.")

# === Prepare Cumulative Output Containers ===
all_violations = []
history_df = (
    pd.read_excel(history_file, index_col=0)
    if os.path.exists(history_file)
    else pd.DataFrame()
)

# === Step 2: Process Each Report ===
for report_path in sorted(report_files, key=os.path.getmtime):
    # Extract date from filename
    date_match = re.search(r"(\d{2}_\d{2}_\d{4})", report_path)
    if not date_match:
        print(f"Skipping file (no date found): {report_path}")
        continue
    report_date = datetime.strptime(date_match.group(1), "%m_%d_%Y").strftime(
        "%Y-%m-%d"
    )

    print(f"Processing report: {os.path.basename(report_path)} ({report_date})")

    # Load workbook
    wb = load_workbook(report_path, data_only=True)
    sheetnames = wb.sheetnames

    # === A: Photo Quality Check ===
    ignore_sheets = {"Summary", "Guide", "Google Shopping Rejections"}
    for sheetname in sheetnames:
        if sheetname in ignore_sheets:
            continue

        ws = wb[sheetname]
        min_photos = None

        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    match = re.search(r"Make sure there (?:are|is) (\d+) photos", cell)
                    if match:
                        min_photos = int(match.group(1))
                        break
            if min_photos is not None:
                break

        if min_photos is None:
            continue

        try:
            df = pd.read_excel(report_path, sheet_name=sheetname, skiprows=43)
        except Exception as e:
            print(f"Error reading sheet '{sheetname}': {e}")
            continue

        df.columns = [str(c).strip() for c in df.columns]
        if not {"Number of photos", "Item Id", "Custom label"}.issubset(df.columns):
            continue

        for _, row in df.iterrows():
            raw_value = str(row["Number of photos"]).strip()
            match = re.search(r"\d+", raw_value)
            if match:
                num_photos = int(match.group())
                custom_label = str(row["Custom label"]).strip()
                if num_photos < min_photos and custom_label.lower() != "not provided":
                    item_id = row["Item Id"]
                    reason = (
                        f"The number of photos ({num_photos}) is less than the minimum "
                        f"number of photos ({min_photos})"
                    )
                    all_violations.append(
                        {
                            "Item Id": item_id,
                            "Custom label": custom_label,
                            "Reason": reason,
                            "Date": report_date,
                        }
                    )

    # === B: Summary Ranking Tracking ===
    if "Summary" in sheetnames:
        summary_sheet = wb["Summary"]
        category_percentiles = {}

        for row in summary_sheet.iter_rows(values_only=True):
            row_str = " ".join([str(cell) for cell in row if cell])
            cat_match = re.search(
                r"(.+?)\s+/\s+Listings condition: (New|Used|Refurbished)", row_str
            )
            rank_match = re.search(
                r"Your rank by sales \(GMV\) value: (\d+) out of ([\d,]+)", row_str
            )

            if cat_match and rank_match:
                category = cat_match.group(1).strip()
                condition = cat_match.group(2)
                rank = int(rank_match.group(1))
                total = int(rank_match.group(2).replace(",", ""))
                percentile = round(100 * (1 - rank / total), 2)
                key = f"{category} ({condition})"
                category_percentiles[key] = percentile

        if category_percentiles:
            if report_date not in history_df.columns:
                history_df[report_date] = pd.NA
            for category, percentile in category_percentiles.items():
                if category not in history_df.index:
                    history_df.loc[category] = pd.NA
                history_df.at[category, report_date] = percentile

# === Step 3: Finalize and Save Outputs ===
# Sort columns by date
if not history_df.empty:
    history_df.columns = sorted(
        history_df.columns, key=lambda x: datetime.strptime(x, "%Y-%m-%d")
    )
    history_df.to_excel(history_file)
    print(f"[✓] Updated category ranking history saved to: {history_file}")

# Save all violations
if all_violations:
    pd.DataFrame(all_violations).to_csv(output_csv, index=False)
    print(f"[✓] Photo quality violations saved to: {output_csv}")
else:
    print("[✓] No photo quality violations found in any reports.")
