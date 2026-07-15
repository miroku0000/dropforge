from openpyxl import load_workbook
from pathlib import Path
import os
import glob
import json
import wmi
import time
# Initializing the wmi constructor



def isRunning(s):
	# Iterating through all the running processes
	f = wmi.WMI()
	count=0
	for process in f.Win32_Process():	
		# Displaying the P_ID and P_Name of the process
		if process.CommandLine and s in process.CommandLine and "python.exe" in process.CommandLine:
			count=count+1
			print(str(process.ProcessId) +" " + process.CommandLine)
			if count>1:
				return True
	return False
	

def getItemSpecificsFromFile(id):
	print("debug getItemSpecificsFromFile called with id = " + id)
	myvars = {}
	with open("D:\\zikprocessor\\data\\itemspecifics\\" + id,  encoding="utf-8", errors="ignore") as myfile:
		for line in myfile:
			name, var = line.partition(":")[::2]		
			myvars[name.strip()] = var.strip()
		# 	print("debug key= |"+ name +"|")
		# 	print("debugvalue= |"+ var +"|")
		return myvars

#wb = load_workbook(filename = 'C:\\Users\\mirok\\Downloads\\eBay-Active-Listings-Item-Specifics-Feb-18-2023-12_19_52-0700-1384784468.xlsx')


def processFile(filename):
	wb = load_workbook(filename = filename)
	ws = wb['Listings']
	wis=wb['Aspects']
	for row in reversed(range(5, ws.max_row+1)):
		id = ws.cell(row, 2).value
		if id:
			category = ws.cell(row, 1).value
			path = Path("D:\\zikprocessor\\data\\itemspecifics\\" + str(id))
			print(path)
			if not path.is_file():
				os.system("getitemspecificsfromebay.py " + id + " 1> D:\\zikprocessor\\data\\itemspecifics\\" + id)
			
if  not isRunning("populateitemspecificscache.py"):
	username = os.getenv("USERNAME")
	filenames = glob.glob("C:\\Users\\" + username +"\\Downloads\\eBay-Active-Listings-Item-Specifics*.xlsx")
	
	while filenames and len(filenames)>0:
		i=1
		print(filenames)
		l=len(filenames)
		filename=filenames.pop()
		print(str(i) + " of " + str(l) +" files :" + filename)
		processFile(filename) 
		i=i+1
		#os.system("exceltocsv.py")
		#os.system("exceltocsv.py")
		l=len(filenames)
		# filenames = glob.glob("C:\\Users\\" + username +"\\eBay-Active-Listings-Item-Specifics*.xlsx")
else:
	print("populateitemspecificcache.py is already running...")
