"""
Refresh the `transactions` tab of the P&L workbook with 2026 PriceYak orders
(actual costs), and restore the COGS SUMIFS formula so `profit` auto-computes.

- Total Fulfillment Cost (col Q) = ACTUAL cost: zinc converted_payment_total for
  auto orders, your noted price for external, $0 for cancelled or genuinely
  refunded (ReturnRequestClosedWithRefund / 'refund' note). No-refund returns
  keep their cost. (Matches pnl_month.cost_of.)
- Order Date (col A) = Pacific date (so the existing SUMIFS month buckets line up).
- Dedup: existing 2026 rows are removed and re-added fresh (corrects stale values
  like the $81.49 force-cancel case). 2022-2025 history is untouched.
- COGS row 9 (cols N..R) is set back to the sheet's SUMIFS formula.

A backup already exists at ...BACKUP.xlsx. Run pnl_build first to sanity-check totals.
"""

import datetime
import requests
import openpyxl
from zoneinfo import ZoneInfo
from openpyxl.utils import get_column_letter as CL

from pnl_month import cost_of

PATH = r"D:\Ebaytracking\Ebay Profit 2025.xlsx"
PAC = ZoneInfo("America/Los_Angeles")
import config
PY_ACCOUNT_ID = config.PY_ACCOUNT_ID
PY_API_KEY = config.PY_API_KEY

# COGS SUMIFS, matching the sheet's existing pattern (col letter substituted).
COGS_FORMULA = ('=SUMIFS(transactions!$Q:$Q, transactions!$A:$A, ">="&{c}1, '
                'transactions!$A:$A, "<="&EOMONTH({c}1, 0), transactions!$A:$A, '
                '">="&DATE(YEAR({c}1),1,1), transactions!$A:$A, "<="&DATE(YEAR({c}1),12,31))')


def py_login():
    r = requests.post(f"https://www.priceyak.com/v0/account/{PY_ACCOUNT_ID}/api_login",
                      json={"api_key": PY_API_KEY}, timeout=30)
    r.raise_for_status()
    return r.json()["token"]


def fetch_year(token, year):
    """All non-cancelled-irrelevant orders created within `year` (Pacific)."""
    jan1 = datetime.datetime(year, 1, 1, tzinfo=PAC).timestamp()
    nextjan = datetime.datetime(year + 1, 1, 1, tzinfo=PAC).timestamp()
    headers = {"Authorization": "Bearer " + token, "Accept": "*/*"}
    orders, off = [], 0
    while off < 6000:
        d = requests.get(f"https://www.priceyak.com/v0/account/{PY_ACCOUNT_ID}/orders",
                         headers=headers, params={"count": 100, "offset": off}, timeout=60).json().get("data", [])
        if not d:
            break
        orders += d
        off += len(d)
        if min((o.get("created_time") or 9e18) for o in d) < jan1:
            break
    return [o for o in orders if jan1 <= (o.get("created_time") or 0) < nextjan]


def _scalar(v):
    """openpyxl only accepts primitives -- collapse anything else to None."""
    return v if isinstance(v, (str, int, float, bool, datetime.datetime)) else None


def order_row(o):
    db = o.get("destination_blob") or {}
    dt = datetime.datetime.fromtimestamp(o["created_time"], PAC)
    order_date = datetime.datetime(dt.year, dt.month, dt.day)  # date-only, matches sheet
    q = 0.0 if o.get("cancelled") else round(cost_of(o)[0], 2)
    item_ids = db.get("itemIds") or []
    fa = o.get("fulfillment_account")
    fa_email = fa.get("username") if isinstance(fa, dict) else fa
    row = [
        order_date,                                    # A Order Date
        o.get("state"),                                # B State
        None,                                          # C Display Status
        o.get("retailer") or "amazon",                 # D Source
        o.get("destination_order_id"),                 # E Destination Order ID
        o.get("destination_record_number"),            # F Destination Record Number
        item_ids[0] if item_ids else None,             # G Destination Item ID
        db.get("txnPrice"),                            # H Txn Price
        db.get("fvfFee"),                              # I FVF Fee
        db.get("amountPaid"),                          # J Amount Paid
        db.get("quantity"),                            # K Order Quantity
        db.get("sku"),                                 # L Source Item ID (ASIN)
        None,                                          # M Shipping Name
        None,                                          # N Shipping Address
        o.get("fulfillment_order_id"),                 # O Source Order IDs
        None,                                          # P Cost of Goods Sold
        q,                                             # Q Total Fulfillment Cost  <-- COGS sums this
        fa_email,                                      # R Fulfillment Account Email
        o.get("fulfillment_tracking_number") or o.get("partner_tracking_number"),  # S Tracking
        round((o.get("profit") or 0) / 100, 2),        # T Profit
        (o.get("frontend_details") or {}).get("orderNotes"),  # U Notes
    ]
    return [_scalar(v) for v in row]


def refresh_year_transactions(tx, year, orders):
    """Replace `year` rows in the transactions sheet with fresh API rows. 2025-
    and-earlier history is untouched. Returns count written."""
    rows_y = [r for r in range(2, tx.max_row + 1)
              if isinstance(tx.cell(r, 1).value, datetime.datetime) and tx.cell(r, 1).value.year == year]
    if rows_y:
        first = min(rows_y)
        tx.delete_rows(first, max(rows_y) - first + 1)
    else:
        data_rows = [r for r in range(2, tx.max_row + 1) if tx.cell(r, 1).value is not None]
        first = (max(data_rows) + 1) if data_rows else 2
    rows = sorted((order_row(o) for o in orders), key=lambda r: r[0])
    date_fmt = tx.cell(2, 1).number_format
    for i, row in enumerate(rows):
        rr = first + i
        for c, val in enumerate(row, 1):
            tx.cell(rr, c).value = val
        tx.cell(rr, 1).number_format = date_fmt
    return len(rows)


def main():
    year = 2026
    token = py_login()
    orders = fetch_year(token, year)
    print(f"Pulled {len(orders)} {year} order(s) from PriceYak.")
    wb = openpyxl.load_workbook(PATH)
    for c in ("N", "O", "P", "Q", "R"):
        wb["profit"][f"{c}9"] = COGS_FORMULA.format(c=c)
        wb["profit"][f"{c}9"].number_format = wb["profit"]["M9"].number_format
    n = refresh_year_transactions(wb["transactions"], year, orders)
    print(f"refreshed {n} {year} transaction row(s); restored COGS formulas")
    wb.save(PATH)
    print(f"Saved {PATH}")


if __name__ == "__main__":
    main()
