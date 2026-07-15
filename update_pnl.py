"""
Update the monthly P&L workbook for every COMPLETE month up through LAST month
(derived from today's date -- never the current partial month), then refresh the
transactions tab so COGS auto-computes. Intended to run from airotate.bat.

Per month it writes on the `profit` tab:
  - Revenue + net sales  (scraped from the eBay Sales dashboard)
  - priceyak  (fixed 156.75/mo)
  - month-header date, and COGS / total Expenses / profit FORMULAS
crawlbase is left for you to fill (usage-based, no API). COGS comes from the
transactions tab, which this refreshes with actual PriceYak costs.

NOTE: edits the LOCAL xlsx. To sync the online Google Sheet, keep the file in a
Drive-synced folder or re-upload it.

Usage:
    python update_pnl.py            # auto: complete months through last month
    python update_pnl.py --no-ebay  # skip the eBay scrape (COGS/transactions only)
"""

import sys
import datetime
import calendar

import openpyxl
from openpyxl.utils import get_column_letter as CL
from zoneinfo import ZoneInfo

from refresh_transactions import (PATH, PAC, COGS_FORMULA, py_login, fetch_year,
                                  refresh_year_transactions)
import crawlbase_cost

PRICEYAK_FEE = 156.75
ROW = {"date": 1, "rev": 2, "net": 3, "py": 7, "cb": 8, "cogs": 9, "total": 12, "profit": 14}


def target_months(today):
    """(year, [months]) for every complete month of last-month's year through
    last month. e.g. 2026-06-12 -> (2026, [1..5]); 2027-01-09 -> (2026, [1..12])."""
    y, m = today.year, today.month
    ly, lm = (y - 1, 12) if m == 1 else (y, m - 1)   # last complete month
    return ly, list(range(1, lm + 1))


def find_or_add_col(ws, year, month):
    for c in range(2, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, datetime.datetime) and v.year == year and v.month == month:
            return c
    last = max(c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value is not None)
    nc = last + 1
    ws.cell(1, nc).value = datetime.datetime(year, month, 1)
    ws.cell(1, nc).number_format = ws.cell(1, last).number_format
    return nc


def _set(ws, row, col, value, fmt_from_col=None):
    cell = ws.cell(row, col)
    cell.value = value
    if fmt_from_col:
        cell.number_format = ws.cell(row, fmt_from_col).number_format


def main():
    no_ebay = "--no-ebay" in sys.argv
    today = datetime.datetime.now(PAC)
    year, months = target_months(today)
    print(f"Today {today:%Y-%m-%d}; updating complete months: {year}-{months[0]:02d}..{year}-{months[-1]:02d}")

    # 1) Refresh the transactions tab for the year (so COGS auto-computes).
    token = py_login()
    orders = fetch_year(token, year)
    wb = openpyxl.load_workbook(PATH)
    pr = wb["profit"]
    n = refresh_year_transactions(wb["transactions"], year, orders)
    print(f"refreshed {n} {year} transaction row(s)")

    # crawlbase: snapshot the current month's usage + cost-per-1000 rate.
    cb_data = crawlbase_cost.record_current()
    print(f"crawlbase rate: ${cb_data.get('rate_per_1000')}/1000")

    # 2) eBay revenue / net sales per complete month (one browser session).
    ebay = {}
    if not no_ebay:
        from playwright.sync_api import sync_playwright
        from playwright_browser import launch_ebay_browser
        import ai_ebay_sales_performance as eb
        with sync_playwright() as p:
            ctx = launch_ebay_browser(p, viewport={"width": 1500, "height": 950}, accept_downloads=False)
            page = ctx.pages[0] if ctx.pages else ctx.new_page()
            try:
                for mo in months:
                    try:
                        ebay[mo] = eb.extract(page, year, mo)
                    except Exception as e:
                        print(f"  eBay scrape {year}-{mo:02d} failed: {str(e)[:70]}")
                        ebay[mo] = {}
            finally:
                ctx.close()

    # 3) Fill each month's column.
    for mo in months:
        c = find_or_add_col(pr, year, mo)
        L = CL(c)
        prev = c - 1  # mirror number formats from the previous month column
        e = ebay.get(mo, {})
        if e.get("revenue") is not None:
            _set(pr, ROW["rev"], c, round(e["revenue"], 2), prev)
        if e.get("net_sales") is not None:
            _set(pr, ROW["net"], c, round(e["net_sales"], 2), prev)
        _set(pr, ROW["py"], c, PRICEYAK_FEE, prev)
        cb = crawlbase_cost.month_cost(cb_data, f"{year}-{mo:02d}")
        if cb is not None:
            _set(pr, ROW["cb"], c, cb, prev)
        _set(pr, ROW["cogs"], c, COGS_FORMULA.format(c=L), prev)
        _set(pr, ROW["total"], c, f"=sum({L}{ROW['py']}:{L}{ROW['cb']+2})", prev)
        _set(pr, ROW["profit"], c, f"={L}{ROW['net']}-{L}{ROW['total']}", prev)
        print(f"  {year}-{mo:02d} -> col {L}: rev={e.get('revenue')} net={e.get('net_sales')} (COGS/total/profit=formula)")

    wb.save(PATH)
    print(f"Saved {PATH}")
    print("Reminder: re-upload the xlsx (or keep it in a Drive-synced folder) to sync the online sheet.")


if __name__ == "__main__":
    main()
